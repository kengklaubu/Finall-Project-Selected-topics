import os
from openpyxl import load_workbook
from django.conf import settings
from easypark.models import CustomUser, ParkingLocation, ParkingSpot, Reservation
from django.core.management.base import BaseCommand
from datetime import datetime
from datetime import datetime

class Command(BaseCommand):
    help = 'Load data from Excel into CustomUser, ParkingLocation, ParkingSpot, and Reservation models'

    def handle(self, *args, **options):
        file_path = os.path.join(settings.BASE_DIR, 'easypark', 'fixtures', 'mockup-data.xlsx')

        # Load the Excel workbook
        wb = load_workbook(file_path)

        
        # Load CustomUser data
        user_sheet = wb['easypark_customuser']
        for row in user_sheet.iter_rows(min_row=2, values_only=True):
    # ตรวจสอบจำนวนคอลัมน์และดึงเฉพาะที่ต้องการ
            row = row[:12]  # ดึงเฉพาะ 12 คอลัมน์แรก
            user_id, last_login, is_superuser, username, first_name, last_name, email, is_staff, is_active, date_joined, role, password = row

            # ตรวจสอบและจัดการค่า username ที่ว่าง
            if not user_id or user_id == "":
                user_id = CustomUser.objects.count() + 1  # ใช้จำนวนผู้ใช้ที่มีอยู่ +1
            if not username or username.strip() == "":
                username = f"user_{user_id}"

            if not first_name or first_name.strip() == "":
                first_name = "DefaultFirstName"

            if not last_name or last_name.strip() == "":
                last_name = "DefaultLastName"

            if not email or email.strip() == "":
                email = f"email_{user_id}@example.com"

            if not date_joined:
                date_joined = datetime.now()

            if not role or role.strip() == "":
                role = "user"

            user_instance, created = CustomUser.objects.get_or_create(
                id=user_id,
                defaults={
                    'username': username,
                    'email': email,
                    'first_name': first_name,
                    'last_name': last_name,
                    'is_staff': bool(is_staff),
                    'is_superuser': bool(is_superuser),
                    'is_active': bool(is_active),
                    'date_joined': date_joined,
                    'last_login': last_login,
                    'role': role,
                }
            )
            if created:
                user_instance.set_password(password)
                user_instance.save()
                self.stdout.write(f"Created user: {username}")
            else:
                self.stdout.write(f"User already exists: {username}")







        location_sheet = wb['easypark_parkinglocation']
        for row in location_sheet.iter_rows(min_row=2, values_only=True):
    # ดึงเฉพาะ 8 คอลัมน์แรก
            row = row[:8]
            location_id, name, slug, description, total_spots, available_spots, created_at, updated_at = row

    # กำหนดค่าเริ่มต้นสำหรับฟิลด์ที่อาจว่าง
            created_at = created_at or datetime.now()
            updated_at = updated_at or datetime.now()

    # บันทึกข้อมูลลงใน ParkingLocation
            ParkingLocation.objects.get_or_create(
                id=location_id,
                defaults={
                    'name': name,
                    'slug': slug,
                    'description': description,
                    'total_spots': total_spots,
                    'available_spots': available_spots,
                    'created_at': created_at,
                    'updated_at': updated_at,
                }
            )



        # Load ParkingSpot data
        

        spot_sheet = wb['easypark_parkingspot']
        for row in spot_sheet.iter_rows(min_row=2, values_only=True):
    # ดึงเฉพาะ 9 คอลัมน์แรก
            row = row[:9]
            spot_id, spot_number, is_available, date, license_plate, created_at, updated_at, location_id, reserved_by_id = row

    # ตรวจสอบ location และ reserved_by
            location = ParkingLocation.objects.filter(id=location_id).first()
            reserved_by = CustomUser.objects.filter(id=reserved_by_id).first() if reserved_by_id else None

    # กำหนดค่าเริ่มต้นสำหรับฟิลด์ที่อาจว่าง
            created_at = created_at or datetime.now()
            updated_at = updated_at or datetime.now()

    # บันทึกข้อมูลลงใน ParkingSpot
            if location:
                ParkingSpot.objects.get_or_create(
                    id=spot_id,
                    defaults={
                        'spot_number': spot_number,
                        'is_available': bool(is_available),
                        'date': date,
                        'license_plate': license_plate,
                        'created_at': created_at,
                        'updated_at': updated_at,
                        'location': location,
                        'reserved_by': reserved_by,
                    }
                )




        reservation_sheet = wb['easypark_reservation']
        for row in reservation_sheet.iter_rows(min_row=2, values_only=True):
    # ดึงเฉพาะ 7 คอลัมน์แรก
            row = row[:7]
            reservation_id, user_id, parking_spot_id, reservation_date, start_time, end_time, status = row

    # ตรวจสอบว่าค่า id เป็นตัวเลข
            if not isinstance(reservation_id, int):
                self.stdout.write(f"Invalid reservation_id: {reservation_id}. Skipping...")
                continue
            if not isinstance(user_id, int):
                self.stdout.write(f"Invalid user_id: {user_id}. Skipping...")
                continue
            if not isinstance(parking_spot_id, int):
                self.stdout.write(f"Invalid parking_spot_id: {parking_spot_id}. Skipping...")
                continue

    # ตรวจสอบผู้ใช้และที่จอดรถ
            user = CustomUser.objects.filter(id=user_id).first()
            parking_spot = ParkingSpot.objects.filter(id=parking_spot_id).first()

            if user and parking_spot:
                # ตรวจสอบและแปลงวันที่
                try:
                    reservation_date = datetime.strptime(str(reservation_date), '%Y-%m-%d').date()
                except ValueError as e:
                    self.stdout.write(f"Invalid date format for reservation ID {reservation_id}. Skipping...")
                    continue

                # ตรวจสอบเวลาเริ่มต้นและสิ้นสุด
                if not isinstance(start_time, datetime.time):
                    self.stdout.write(f"Invalid start_time: {start_time}. Skipping...")
                    continue
                if not isinstance(end_time, datetime.time):
                    self.stdout.write(f"Invalid end_time: {end_time}. Skipping...")
                    continue

                # สร้างหรืออัปเดตข้อมูลการจอง
                Reservation.objects.get_or_create(
                    id=reservation_id,
                    defaults={
                        'user': user,
                        'parking_spot': parking_spot,
                        'reservation_date': reservation_date,
                        'reservation_start_time': start_time,
                        'reservation_end_time': end_time,
                        'status': status,
                    }
                )




        self.stdout.write(self.style.SUCCESS("Data loaded successfully!"))
